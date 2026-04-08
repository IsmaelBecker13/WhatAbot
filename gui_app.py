import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import xlrd
from typing import List, Dict, Optional
import os
import requests
import time
import threading
from datetime import datetime, timedelta
from urllib.parse import quote
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException


class WhatsAppMensajeApp:
    def __init__(self, root):
        self.root = root
        self.root.title("WhatsApp Renovaciones - Gestor de Mensajes")
        self.root.geometry("1200x800")
        self.root.configure(bg="#f0f0f0")
        
        # Variables
        self.datos_excel = []
        self.archivo_cargado = ""
        self.columnas_disponibles = []
        self.enviar_hilo = None
        self.detener_envio = threading.Event()
        self.pausar_envio = threading.Event()
        self.pausar_envio.set()
        self.ventana_progreso_envio = None
        self.btn_enviar = None
        self.whatsapp_driver = None
        
        # Crear interfaz
        self.crear_interfaz()
    
    def crear_interfaz(self):
        """Crear elementos de la interfaz gráfica"""
        
        # ===== PANEL SUPERIOR (Carga de archivo) =====
        panel_superior = ttk.Frame(self.root)
        panel_superior.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(panel_superior, text="Archivo Excel:").pack(side=tk.LEFT, padx=5)
        self.label_archivo = ttk.Label(panel_superior, text="Ninguno cargado", 
                                       foreground="gray")
        self.label_archivo.pack(side=tk.LEFT, padx=5)
        
        btn_cargar = ttk.Button(panel_superior, text="Cargar Archivo", 
                               command=self.cargar_archivo)
        btn_cargar.pack(side=tk.LEFT, padx=5)
        
        # ===== PANEL CENTRAL (Dos columnas) =====
        panel_central = ttk.Frame(self.root)
        panel_central.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # --- COLUMNA IZQUIERDA (Columnas disponibles) ---
        frame_izq = ttk.LabelFrame(panel_central, text="Columnas Disponibles", padding=10)
        frame_izq.pack(side=tk.LEFT, fill=tk.BOTH, expand=False, padx=5)
        
        # Info sobre columnas
        ttk.Label(frame_izq, text="Haz clic para copiar:", 
                 font=("Arial", 9, "italic")).pack(anchor=tk.W)
        
        # Listbox de columnas
        scrollbar_col = ttk.Scrollbar(frame_izq)
        scrollbar_col.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.listbox_columnas = tk.Listbox(frame_izq, height=20, width=25,
                                           yscrollcommand=scrollbar_col.set)
        self.listbox_columnas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_col.config(command=self.listbox_columnas.yview)
        
        # Bind para copiar al hacer clic
        self.listbox_columnas.bind('<<ListboxSelect>>', self.copiar_columna)
        
        ttk.Label(frame_izq, text="(clic = copiar a plantilla)", 
                 font=("Arial", 8, "italic"), foreground="gray").pack(anchor=tk.W)
        
        # --- COLUMNA DERECHA (Plantilla y vista previa) ---
        frame_der = ttk.Frame(panel_central)
        frame_der.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5)
        
        # --- Plantilla de mensaje ---
        ttk.Label(frame_der, text="Plantilla de Mensaje", 
                 font=("Arial", 10, "bold")).pack(anchor=tk.W)
        
        ttk.Label(frame_der, text="Usa {NombreColumna} para variables", 
                 font=("Arial", 9, "italic"), foreground="gray").pack(anchor=tk.W)
        
        frame_texto = ttk.Frame(frame_der)
        frame_texto.pack(fill=tk.BOTH, expand=True, pady=5)
        
        scrollbar_texto = ttk.Scrollbar(frame_texto)
        scrollbar_texto.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.texto_plantilla = tk.Text(frame_texto, height=15, wrap=tk.WORD,
                                       yscrollcommand=scrollbar_texto.set)
        self.texto_plantilla.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_texto.config(command=self.texto_plantilla.yview)
        
        # Texto por defecto
        plantilla_default = """Hola {APELLIDO},

Nos comunicamos en relación a la póliza del dominio {DOMINIO}, que el día {FECHAVENCIMIENTOCUOTA} termina su vigencia.

Si desea que se le renueve la vigencia, responda por favor este mensaje afirmando esta opción.

Saludos,
Becker Cristian"""
        self.texto_plantilla.insert(1.0, plantilla_default)
        
        # --- Vista previa ---
        ttk.Label(frame_der, text="Vista Previa (Primer Registro)", 
                 font=("Arial", 10, "bold")).pack(anchor=tk.W, pady=(10, 0))
        
        frame_preview = ttk.Frame(frame_der)
        frame_preview.pack(fill=tk.BOTH, expand=True, pady=5)
        
        scrollbar_preview = ttk.Scrollbar(frame_preview)
        scrollbar_preview.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.texto_preview = tk.Text(frame_preview, height=10, wrap=tk.WORD,
                                     yscrollcommand=scrollbar_preview.set,
                                     state=tk.DISABLED, bg="#e8f5e9")
        self.texto_preview.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_preview.config(command=self.texto_preview.yview)
        
        # ===== PANEL INFERIOR (Botones de acción) =====
        panel_inferior = ttk.Frame(self.root)
        panel_inferior.pack(fill=tk.X, padx=10, pady=10)
        
        btn_actualizar = ttk.Button(panel_inferior, text="Actualizar Vista Previa",
                                   command=self.actualizar_preview)
        btn_actualizar.pack(side=tk.LEFT, padx=5)
        
        self.btn_enviar = ttk.Button(panel_inferior, text="Enviar Mensajes por WhatsApp",
                          command=self.enviar_mensajes)
        self.btn_enviar.pack(side=tk.LEFT, padx=5)
        
        # Separador visual
        ttk.Label(panel_inferior, text="  |  ").pack(side=tk.LEFT, padx=2)
        
        btn_cargar_plantilla = ttk.Button(panel_inferior, text="Cargar Plantilla",
                                         command=self.cargar_plantilla)
        btn_cargar_plantilla.pack(side=tk.LEFT, padx=5)
        
        btn_guardar = ttk.Button(panel_inferior, text="Guardar Plantilla",
                                command=self.guardar_plantilla)
        btn_guardar.pack(side=tk.LEFT, padx=5)
        
        # Etiqueta de estado
        self.label_estado = ttk.Label(panel_inferior, text="Listo", foreground="green")
        self.label_estado.pack(side=tk.RIGHT, padx=5)
    
    def cargar_archivo(self):
        """Cargar archivo Excel (.xls o .xlsx)"""
        archivo = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel files", "*.xls *.xlsx"), ("All files", "*.*")]
        )
        
        if not archivo:
            return
        
        self.procesar_archivo(archivo)
    
    def procesar_archivo(self, archivo: str):
        """Procesar y cargar un archivo Excel"""
        print(f"🔄 Iniciando carga de: {archivo}")
        try:
            self.archivo_cargado = archivo
            
            # Determinar formato y cargar
            if archivo.endswith('.xls'):
                print("📄 Detectado formato .xls (antiguo)")
                # Archivo Excel antiguo (.xls)
                wb = xlrd.open_workbook(archivo)
                ws = wb.sheet_by_index(0)
                print(f"✓ Archivo abierto. Filas: {ws.nrows}, Columnas: {ws.ncols}")
                
                # Obtener columnas (primera fila)
                self.columnas_disponibles = []
                for col_idx in range(ws.ncols):
                    cell = ws.cell(0, col_idx)
                    valor = cell.value
                    valor_str = str(valor).strip() if valor else f"Columna{col_idx}"
                    self.columnas_disponibles.append(valor_str)
                print(f"✓ Columnas encontradas: {self.columnas_disponibles}")
                
                # Obtener datos
                self.datos_excel = []
                for row_idx in range(1, ws.nrows):
                    fila_dict = {}
                    tiene_datos = False
                    for col_idx, col_nombre in enumerate(self.columnas_disponibles):
                        valor = ws.cell_value(row_idx, col_idx)
                        fila_dict[col_nombre] = valor
                        if valor:
                            tiene_datos = True
                    
                    if tiene_datos:
                        self.datos_excel.append(fila_dict)
                
                print(f"✓ Datos cargados: {len(self.datos_excel)} filas")
            else:
                print("📄 Detectado formato .xlsx (moderno)")
                # Archivo Excel nuevo (.xlsx)
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active
                
                # Obtener columnas
                self.columnas_disponibles = []
                for cell in ws[1]:
                    if cell.value:
                        self.columnas_disponibles.append(str(cell.value).strip())
                print(f"✓ Columnas encontradas: {self.columnas_disponibles}")
                
                # Obtener datos
                self.datos_excel = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(row):
                        fila_dict = {}
                        for i, col in enumerate(self.columnas_disponibles):
                            fila_dict[col] = row[i] if i < len(row) else ""
                        self.datos_excel.append(fila_dict)
                
                print(f"✓ Datos cargados: {len(self.datos_excel)} filas")
            
            # Actualizar UI
            self.label_archivo.config(
                text=f"{os.path.basename(archivo)} ({len(self.datos_excel)} registros)",
                foreground="green"
            )
            
            # Mostrar columnas
            self.listbox_columnas.delete(0, tk.END)
            for col in self.columnas_disponibles:
                self.listbox_columnas.insert(tk.END, f"{{{col}}}")
            
            self.label_estado.config(text=f"Archivo cargado: {len(self.datos_excel)} registros")
            print("✓ Actualizando vista previa...")
            self.actualizar_preview()
            print("✅ Archivo cargado exitosamente!")
            
        except Exception as e:
            print(f"❌ ERROR: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Error", f"Error al cargar archivo:\n{str(e)}")
            self.label_estado.config(text="Error al cargar archivo", foreground="red")
    
    def copiar_columna(self, event=None):
        """Copiar nombre de columna a la plantilla"""
        try:
            seleccion = self.listbox_columnas.curselection()
            if seleccion:
                texto = self.listbox_columnas.get(seleccion[0])
                self.texto_plantilla.insert(tk.INSERT, texto)
        except Exception as e:
            messagebox.showerror("Error", f"Error: {str(e)}")
    
    def actualizar_preview(self):
        """Actualizar vista previa con el primer registro"""
        if not self.datos_excel:
            messagebox.showwarning("Advertencia", "Carga un archivo Excel primero")
            return
        
        try:
            plantilla = self.texto_plantilla.get(1.0, tk.END)
            mensaje = self.generar_mensaje(plantilla, self.datos_excel[0])
            
            self.texto_preview.config(state=tk.NORMAL)
            self.texto_preview.delete(1.0, tk.END)
            self.texto_preview.insert(1.0, mensaje)
            self.texto_preview.config(state=tk.DISABLED)
            
            self.label_estado.config(text="Vista previa actualizada", foreground="blue")
        
        except Exception as e:
            messagebox.showerror("Error", f"Error al generar vista previa:\n{str(e)}")
            self.label_estado.config(text="Error en vista previa", foreground="red")
    
    def convertir_fecha_excel(self, valor):
        """Convertir número de fecha de Excel a formato dd/mm/yyyy"""
        # Fecha base de Excel (1900-01-01, pero Excel cuenta desde 1899-12-30)
        fecha_base = datetime(1899, 12, 30)
        
        try:
            # Si es un número (float o int), interpretarlo como días desde la fecha base
            if isinstance(valor, (int, float)):
                dias = int(valor)
                fecha = fecha_base + timedelta(days=dias)
                return fecha.strftime('%d/%m/%Y')
        except:
            pass
        
        # Si no se puede convertir, retornar como está
        return str(valor).strip()
    
    def limpiar_telefono(self, telefono: str) -> str:
        """Limpiar y formatear número de teléfono para Argentina"""
        if not telefono:
            return ""
        
        # DEBUG: Ver exactamente qué llega
        print(f"  DEBUG limpiar_telefono - Entrada: '{telefono}' (tipo: {type(telefono).__name__})")
        
        # Si es float, convertir a int primero para eliminar .0
        if isinstance(telefono, float):
            telefono = str(int(telefono))
            print(f"  DEBUG - Era float, convertido a int: '{telefono}'")
        else:
            telefono = str(telefono).strip()
        
        print(f"  DEBUG - Después conversión: '{telefono}'")
        
        # Remover puntos, comas y espacios (pero no el punto decimal que ya fue tratado)
        telefono = telefono.replace(',', '').replace(' ', '').replace('.', '')
        print(f"  DEBUG - Después limpiar caracteres: '{telefono}'")
        
        # Si ya tiene el formato correcto +54
        if telefono.startswith('+54'):
            print(f"  DEBUG - Ya tiene +54, retornando: '{telefono}'")
            return telefono
        
        # Si tiene código de país sin +
        if telefono.startswith('54') and len(telefono) >= 12:
            resultado = f"+{telefono}"
            print(f"  DEBUG - Tiene 54 sin +, retornando: '{resultado}'")
            return resultado
        
        # Si empieza con 0 (formato local Argentina: 0223456789)
        if telefono.startswith('0'):
            telefono = telefono[1:]  # Remover el 0
            print(f"  DEBUG - Removió 0 inicial, ahora: '{telefono}'")
        
        # Agregar código de país
        resultado = f"+54{telefono}"
        print(f"  DEBUG - Resultado final: '{resultado}'")
        return resultado
    
    def generar_mensaje(self, plantilla: str, registro: Dict) -> str:
        """Generar mensaje reemplazando variables"""
        mensaje = plantilla
        
        for columna, valor in registro.items():
            # Reemplazar {Columna} con el valor
            placeholder = f"{{{columna}}}"
            
            # Convertir valor a string, manejando fechas y números
            if isinstance(valor, (int, float)):
                # Detectar si es una fecha de Excel (número entre 30000 y 50000)
                if 30000 <= valor <= 50000:
                    valor_str = self.convertir_fecha_excel(valor)
                else:
                    valor_str = str(int(valor)) if valor == int(valor) else str(valor)
            elif hasattr(valor, 'strftime'):
                valor_str = valor.strftime('%d/%m/%Y')
            else:
                valor_str = str(valor).strip() if valor else ""
            
            mensaje = mensaje.replace(placeholder, valor_str)
        
        return mensaje
    
    def enviar_mensajes(self):
        """Enviar mensajes por WhatsApp Web"""
        if not self.datos_excel:
            messagebox.showwarning("Advertencia", "Carga un archivo Excel primero")
            return

        if self.enviar_hilo and self.enviar_hilo.is_alive():
            messagebox.showinfo("Envío en curso", "Ya hay un envío en progreso.")
            return
        
        # Confirmar
        respuesta = messagebox.askyesno(
            "Confirmar",
            f"¿Enviar {len(self.datos_excel)} mensajes por WhatsApp?\n\n"
            "Se abrirá WhatsApp Web en tu navegador.\n"
            "Debes tener WhatsApp Web abierto y sincronizado."
        )
        
        if not respuesta:
            return

        self.detener_envio.clear()
        self.pausar_envio.set()
        if self.btn_enviar:
            self.btn_enviar.config(state=tk.DISABLED)
        
        # Crear ventana de progreso
        ventana_progreso = tk.Toplevel(self.root)
        ventana_progreso.title("Enviando mensajes...")
        ventana_progreso.geometry("500x350")
        ventana_progreso.protocol("WM_DELETE_WINDOW", self.detener_envio_envio)
        self.ventana_progreso_envio = ventana_progreso
        
        frame_progreso = ttk.Frame(ventana_progreso, padding=10)
        frame_progreso.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame_progreso, text="Enviando mensajes por WhatsApp Web...", 
                 font=("Arial", 12, "bold")).pack(anchor=tk.W)
        
        ttk.Label(frame_progreso, text="(No cierres la ventana del navegador)", 
                 font=("Arial", 9, "italic"), foreground="blue").pack(anchor=tk.W)

        frame_controles = ttk.Frame(frame_progreso)
        frame_controles.pack(fill=tk.X, pady=(5, 0))

        self.btn_pausar_envio = ttk.Button(
            frame_controles,
            text="Pausar",
            command=self.toggle_pausa_envio
        )
        self.btn_pausar_envio.pack(side=tk.LEFT, padx=(0, 5))

        self.btn_detener_envio = ttk.Button(
            frame_controles,
            text="Terminar",
            command=self.detener_envio_envio
        )
        self.btn_detener_envio.pack(side=tk.LEFT)
        
        barra_progreso = ttk.Progressbar(frame_progreso, maximum=len(self.datos_excel))
        barra_progreso.pack(fill=tk.X, pady=10)
        
        text_log = tk.Text(frame_progreso, height=14, wrap=tk.WORD)
        text_log.pack(fill=tk.BOTH, expand=True)
        
        plantilla = self.texto_plantilla.get(1.0, tk.END)
        text_log.config(state=tk.DISABLED)

        self.enviar_hilo = threading.Thread(
            target=self._enviar_mensajes_en_segundo_plano,
            args=(plantilla, text_log, barra_progreso),
            daemon=True
        )
        self.enviar_hilo.start()

    def _actualizar_log_envio(self, text_log, mensaje: str):
        self.root.after(0, self._actualizar_log_envio_ui, text_log, mensaje)

    def _actualizar_log_envio_ui(self, text_log, mensaje: str):
        if not text_log.winfo_exists():
            return
        text_log.config(state=tk.NORMAL)
        text_log.insert(tk.END, mensaje)
        text_log.see(tk.END)
        text_log.config(state=tk.DISABLED)

    def _actualizar_progreso_envio(self, barra_progreso, valor: int):
        def actualizar():
            if barra_progreso.winfo_exists():
                barra_progreso.config(value=valor)

        self.root.after(0, actualizar)

    def _esperar_con_control(self, segundos: float) -> bool:
        fin = time.time() + segundos
        while time.time() < fin:
            if self.detener_envio.is_set():
                return False
            if not self.pausar_envio.wait(timeout=0.25):
                continue
            restante = fin - time.time()
            if restante > 0:
                time.sleep(min(0.5, restante))
        return not self.detener_envio.is_set()

    def _finalizar_envio_ui(self, ventana_progreso, enviados: int, errores: int, detenido: bool):
        if ventana_progreso and ventana_progreso.winfo_exists():
            ventana_progreso.destroy()

        if self.btn_enviar:
            self.btn_enviar.config(state=tk.NORMAL)

        self.enviar_hilo = None
        self.ventana_progreso_envio = None

        if detenido:
            self.label_estado.config(text=f"Envío detenido. Enviados: {enviados}, Errores: {errores}", foreground="orange")
        else:
            self.label_estado.config(
                text=f"Enviados: {enviados}, Errores: {errores}",
                foreground="green" if errores == 0 else "orange"
            )

    def _obtener_driver_whatsapp(self):
        if self.whatsapp_driver:
            return self.whatsapp_driver

        try:
            opciones = webdriver.ChromeOptions()
            opciones.add_argument("--disable-notifications")
            self.whatsapp_driver = webdriver.Chrome(options=opciones)
            self.whatsapp_driver.get("https://web.whatsapp.com")

            # Espera a que el usuario escanee QR si hace falta.
            WebDriverWait(self.whatsapp_driver, 120).until(
                EC.presence_of_element_located((By.ID, "pane-side"))
            )
            return self.whatsapp_driver
        except Exception as e:
            print(f"Error iniciando WhatsApp Web con Selenium: {e}")
            self._cerrar_driver_whatsapp()
            return None

    def _cerrar_driver_whatsapp(self):
        if self.whatsapp_driver:
            try:
                self.whatsapp_driver.quit()
            except Exception:
                pass
            finally:
                self.whatsapp_driver = None

    def _enviar_mensajes_en_segundo_plano(self, plantilla: str, text_log, barra_progreso):
        enviados = 0
        errores = 0
        espera_entre_mensajes = 15

        try:
            for i, registro in enumerate(self.datos_excel, 1):
                if self.detener_envio.is_set():
                    break

                while not self.pausar_envio.is_set():
                    if self.detener_envio.is_set():
                        break
                    time.sleep(0.2)

                if self.detener_envio.is_set():
                    break

                try:
                    mensaje = self.generar_mensaje(plantilla, registro)

                    telefono = (registro.get('TELEFONO', '') or 
                              registro.get('Teléfono', '') or
                              registro.get('telefono', '') or
                              registro.get('Celular', '') or
                              registro.get('celular', '') or '')

                    telefono = self.limpiar_telefono(telefono)

                    if not telefono:
                        self._actualizar_log_envio(text_log, f"[{i}] ✗ Teléfono no encontrado\n")
                        errores += 1
                        self._actualizar_progreso_envio(barra_progreso, i)
                        continue

                    if self.enviar_whatsapp_web(telefono, mensaje):
                        self._actualizar_log_envio(text_log, f"[{i}] ✓ Enviado a {telefono}\n")
                        enviados += 1
                    else:
                        self._actualizar_log_envio(text_log, f"[{i}] ✗ Error al enviar a {telefono}\n")
                        errores += 1

                    self._actualizar_progreso_envio(barra_progreso, i)

                    if not self._esperar_con_control(espera_entre_mensajes):
                        break

                except Exception as e:
                    self._actualizar_log_envio(text_log, f"[{i}] ✗ Error: {str(e)}\n")
                    errores += 1
                    self._actualizar_progreso_envio(barra_progreso, i)
        finally:
            self._cerrar_driver_whatsapp()
            self.root.after(0, self._actualizar_log_envio_ui, text_log, f"\n{'='*40}\nEnviados: {enviados}\nErrores: {errores}\n")
            self.root.after(0, self._finalizar_envio_ui, self.ventana_progreso_envio, enviados, errores, self.detener_envio.is_set())
    
    def enviar_whatsapp_web(self, telefono: str, mensaje: str) -> bool:
        """Enviar mensaje por WhatsApp Web usando Selenium"""
        try:
            driver = self._obtener_driver_whatsapp()
            if not driver:
                return False

            telefono_limpio = telefono.replace("+", "").strip()
            url = f"https://web.whatsapp.com/send?phone={telefono_limpio}&text={quote(mensaje)}"
            driver.get(url)

            boton_enviar = WebDriverWait(driver, 35).until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//button[@data-icon='send' or @aria-label='Enviar' or @aria-label='Send']")
                )
            )

            try:
                boton_enviar.click()
            except WebDriverException:
                driver.execute_script("arguments[0].click();", boton_enviar)

            time.sleep(2)
            return True
        except TimeoutException:
            print(f"Timeout esperando botón Enviar para {telefono}")
            return False
        
        except Exception as e:
            print(f"Error enviando WhatsApp con Selenium a {telefono}: {e}")
            return False

    def toggle_pausa_envio(self):
        """Pausar o reanudar el envío en curso"""
        if not self.enviar_hilo or not self.enviar_hilo.is_alive():
            return

        if self.pausar_envio.is_set():
            self.pausar_envio.clear()
            if hasattr(self, 'btn_pausar_envio'):
                self.btn_pausar_envio.config(text="Reanudar")
            if self.label_estado.winfo_exists():
                self.label_estado.config(text="Envío en pausa", foreground="blue")
        else:
            self.pausar_envio.set()
            if hasattr(self, 'btn_pausar_envio'):
                self.btn_pausar_envio.config(text="Pausar")
            if self.label_estado.winfo_exists():
                self.label_estado.config(text="Envío reanudado", foreground="green")

    def detener_envio_envio(self):
        """Detener el envío en curso"""
        self.detener_envio.set()
        self.pausar_envio.set()
        if hasattr(self, 'btn_pausar_envio'):
            self.btn_pausar_envio.config(state=tk.DISABLED)
        if hasattr(self, 'btn_detener_envio'):
            self.btn_detener_envio.config(state=tk.DISABLED)
        if self.label_estado.winfo_exists():
            self.label_estado.config(text="Deteniendo envío...", foreground="orange")
    
    def guardar_plantilla(self):
        """Guardar plantilla a archivo"""
        try:
            archivo = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
            )
            
            if archivo:
                contenido = self.texto_plantilla.get(1.0, tk.END)
                with open(archivo, 'w', encoding='utf-8') as f:
                    f.write(contenido)
                
                messagebox.showinfo("Éxito", f"Plantilla guardada en:\n{archivo}")
                self.label_estado.config(text="Plantilla guardada", foreground="green")
        
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar:\n{str(e)}")
    
    def cargar_plantilla(self):
        """Cargar plantilla desde archivo"""
        try:
            archivo = filedialog.askopenfilename(
                title="Seleccionar plantilla",
                filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
            )
            
            if archivo:
                with open(archivo, 'r', encoding='utf-8') as f:
                    contenido = f.read()
                
                self.texto_plantilla.delete(1.0, tk.END)
                self.texto_plantilla.insert(1.0, contenido)
                
                messagebox.showinfo("Éxito", f"Plantilla cargada desde:\n{archivo}")
                self.label_estado.config(text="Plantilla cargada", foreground="green")
                self.actualizar_preview()
        
        except Exception as e:
            messagebox.showerror("Error", f"Error al cargar:\n{str(e)}")


def main():
    root = tk.Tk()
    app = WhatsAppMensajeApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
